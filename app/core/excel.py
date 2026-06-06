# app/core/excel.py

import openpyxl
import pandas as pd
import re
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# Importiamo le funzioni di utilità che abbiamo creato nel passaggio precedente
from app.core.helpers import is_cosiam, get_color_for_giustificativo

# ===== EXCEL CREATION - STANDARD =====
def create_excel_from_conversion_data(df_out, all_days, names_list, orario_pianificato_data, month_year_header, day_headers, giustificativi_pagati_set, company=None):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Timesheet'
    color_header_bg = '366092'
    color_emp_bg = 'EBF1FA'
    color_white = 'FFFFFF'
    outer_border = Side(style='thin', color='B0B0B0')
    dotted_border = Side(style='dotted', color='A0A0A0')
    thin_border = Side(style='thin', color='D0D0D0')
    font_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    font_emp = Font(name='Calibri', size=11, bold=True, color='000000')
    font_proj = Font(name='Calibri', size=9, color='555555')
    fill_header = PatternFill(start_color=color_header_bg, end_color=color_header_bg, fill_type='solid')
    fill_emp = PatternFill(start_color=color_emp_bg, end_color=color_emp_bg, fill_type='solid')
    fill_white = PatternFill(start_color=color_white, end_color=color_white, fill_type='solid')

    # Flag Cosiam: aggiunge colonna "Ore Ordinarie" tra "TOTALE ORE LAVORATE" e "Straordinari"
    cosiam_flag = is_cosiam(company)

    if cosiam_flag:
        headers = [month_year_header] + day_headers + ['', '', 'TOTALE ORE LAVORATE', 'Ore Ordinarie', 'Straordinari', 'Note']
    else:
        headers = [month_year_header] + day_headers + ['', '', 'TOTALE ORE LAVORATE', 'Straordinari', 'Note']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        if header != '':
            cell.font = font_header
            cell.fill = fill_header
            cell.border = Border(left=thin_border, right=thin_border, top=outer_border, bottom=outer_border)

    # Calcolo offset colonne finali in base alla presenza/assenza della colonna "Ore Ordinarie"
    base_offset = len(all_days)
    col_gap_1 = base_offset + 2
    col_gap_2 = base_offset + 3
    col_totale = base_offset + 4
    if cosiam_flag:
        col_ordinarie = base_offset + 5
        col_straord = base_offset + 6
        col_note = base_offset + 7
    else:
        col_ordinarie = None
        col_straord = base_offset + 5
        col_note = base_offset + 6

    # Tiene traccia delle righe progetti vuote per applicare il merge "N/A" alla fine
    cosiam_empty_project_rows = [] 

    row_idx = 2
    processed_indices = set()
    for i, name_key in enumerate(names_list):
        if i in processed_indices: continue
        is_project_row = name_key.endswith("__PROGETTI")
        has_project_child = False
        if not is_project_row:
            if i + 1 < len(names_list) and names_list[i+1] == f"{name_key}__PROGETTI": has_project_child = True
        
        if is_project_row:
            current_font, current_fill = font_proj, fill_white
            border_top, border_bottom = dotted_border, outer_border
        else:
            current_font, current_fill = font_emp, fill_emp
            border_top = outer_border
            border_bottom = dotted_border if has_project_child else outer_border
        
        current_border = Border(left=thin_border, right=thin_border, top=border_top, bottom=border_bottom)

        # Verifica se questa riga progetti è completamente vuota (solo per Cosiam)
        is_empty_project_row = False
        if cosiam_flag and is_project_row:
            has_any_value = False
            for day_code in all_days:
                v = df_out.loc[name_key, day_code]
                if v is not None and not pd.isna(v) and str(v).strip() != '' and str(v).strip() != '0':
                    has_any_value = True
                    break
            if not has_any_value:
                is_empty_project_row = True

        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            is_gap_column = (col_idx == col_gap_1 or col_idx == col_gap_2)
            
            if is_gap_column: continue

            cell.border = current_border
            cell.fill = current_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            
            if col_idx == 1:
                val = df_out.loc[name_key, 'DIPENDENTI']
                if is_project_row:
                    cell.value = "↳ Progetti"
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.font = Font(name='Calibri', size=9, italic=True, color='666666')
                else:
                    cell.value = val
                    cell.font = current_font
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            elif col_idx <= len(all_days) + 1:
                if is_empty_project_row:
                    pass
                else:
                    valore = df_out.loc[name_key, all_days[col_idx - 2]]
                    cell.value = valore
                    cell.font = current_font
                    if not is_project_row and isinstance(valore, str):
                        if valore == 'FES':
                            cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                            cell.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
                        else:
                            match = re.search(r'([A-Z0-9]{2,4})', valore)
                            if match:
                                code = match.group(1)
                                if code in giustificativi_pagati_set:
                                    color = get_color_for_giustificativo(code)
                                    cell.font = Font(name='Calibri', size=11, color=color, bold=True)
            elif col_idx == col_totale:
                cell.value = df_out.loc[name_key, 'TOTALE ORE LAVORATE']
                cell.font = font_emp if not is_project_row else font_proj
            elif cosiam_flag and col_idx == col_ordinarie:
                cell.value = df_out.loc[name_key, 'Ore Ordinarie']
                cell.font = font_emp if not is_project_row else font_proj
            elif col_idx == col_straord:
                cell.value = df_out.loc[name_key, 'Straordinari']
                cell.font = font_emp if not is_project_row else font_proj
            elif col_idx == col_note:
                cell.value = df_out.loc[name_key, 'Note']
                cell.font = font_emp if not is_project_row else font_proj
        
        if is_empty_project_row:
            cosiam_empty_project_rows.append(row_idx)

        if cosiam_flag:
            ws.row_dimensions[row_idx].height = 22
        
        row_idx += 1
        processed_indices.add(i)

    # Applica i merge "N/A" sulle righe progetti vuote di Cosiam
    if cosiam_flag and cosiam_empty_project_rows:
        first_day_col = 2
        last_day_col = len(all_days) + 1
        start_letter = openpyxl.utils.get_column_letter(first_day_col)
        end_letter = openpyxl.utils.get_column_letter(last_day_col)
        for r in cosiam_empty_project_rows:
            try:
                ws.merge_cells(f'{start_letter}{r}:{end_letter}{r}')
                merged_cell = ws.cell(row=r, column=first_day_col)
                merged_cell.value = "N/A"
                merged_cell.alignment = Alignment(horizontal='center', vertical='center')
                merged_cell.font = Font(name='Calibri', size=10, italic=True, color='999999')
                merged_cell.fill = fill_white
            except Exception:
                pass

    # Dimensioni colonne
    if cosiam_flag:
        ws.column_dimensions['A'].width = 30
        for i, _ in enumerate(all_days, start=2): 
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 12
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_gap_1)].width = 3
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_gap_2)].width = 3
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_totale)].width = 20
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_ordinarie)].width = 20
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_straord)].width = 20
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_note)].width = 20
    else:
        ws.column_dimensions['A'].width = 25
        for i, _ in enumerate(all_days, start=2): 
            ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = 9
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_gap_1)].width = 3
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_gap_2)].width = 3
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_totale)].width = 22
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_straord)].width = 18
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_note)].width = 25
    
    ws.row_dimensions[1].height = 40
    ws.freeze_panes = 'B2'
    return wb

# ===== EXCEL CREATION - SELENA =====
def create_selena_excel(df_out, all_days, names_list, month_year_header, day_headers, giustificativi_pagati_set):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Timesheet'
    ws.sheet_view.showGridLines = False

    font_month = Font(name='Calibri', size=14, bold=True)
    font_emp = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    font_day = Font(name='Calibri', size=11, bold=True)
    font_normal = Font(name='Calibri', size=11)
    
    fill_emp = PatternFill(start_color='366092', end_color='366092', fill_type='solid') 
    fill_zebra_1 = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid') 
    fill_zebra_2 = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid') 
    
    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border_thin = Side(style='thin', color='B0B0B0')
    border_all = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    ws.merge_cells('A1:B1')
    ws['A1'] = month_year_header
    ws['A1'].font = font_month
    ws['A1'].alignment = align_center

    current_row = 3
    num_days = len(all_days)
    real_names = [n for n in names_list if not n.endswith("__PROGETTI")]

    for name in real_names:
        has_projects = f"{name}__PROGETTI" in df_out.index

        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row, end_column=num_days + 1)
        cell_name = ws.cell(row=current_row, column=2, value=name)
        cell_name.font = font_emp
        cell_name.fill = fill_emp
        cell_name.alignment = align_center

        for c in range(2, num_days + 2):
            ws.cell(row=current_row, column=c).border = border_all
        
        current_row += 1

        for col_idx, day_str in enumerate(day_headers, start=2):
            c_day = ws.cell(row=current_row, column=col_idx, value=day_str)
            c_day.font = font_day
            c_day.fill = fill_zebra_1
            c_day.alignment = align_center
            c_day.border = border_all
        
        current_row += 1

        for col_idx, day_code in enumerate(all_days, start=2):
            val = df_out.loc[name, day_code]
            if pd.isna(val): val = ""
            c_val = ws.cell(row=current_row, column=col_idx, value=val)
            c_val.fill = fill_zebra_2
            c_val.alignment = align_center
            c_val.border = border_all
            
            if isinstance(val, str):
                if val == 'FES':
                    c_val.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
                    c_val.font = Font(name='Calibri', size=11, color='FFFFFF', bold=True)
                else:
                    c_val.font = font_normal
                    match = re.search(r'([A-Z0-9]{2,4})', val)
                    if match:
                        code = match.group(1)
                        if code in giustificativi_pagati_set:
                            color = get_color_for_giustificativo(code)
                            selena_custom_colors = {
                                'MP': 'CD5C5C', 'IF': 'FF8C00', '104H': '20B2AA', 
                                '104G': '008080', 'EF': '6B8E23', 'MO': 'DA70D6', 
                                'PO': '6495ED', 'CM': '9932CC', 'CA': '708090'
                            }
                            if code in selena_custom_colors:
                                color = selena_custom_colors[code]
                            c_val.font = Font(name='Calibri', size=11, color=color, bold=True)
            else:
                c_val.font = font_normal
        
        current_row += 1

        if has_projects:
            proj_name = f"{name}__PROGETTI"
            for col_idx, day_code in enumerate(all_days, start=2):
                pval = df_out.loc[proj_name, day_code]
                if pd.isna(pval): pval = ""
                c_pval = ws.cell(row=current_row, column=col_idx, value=pval)
                c_pval.fill = fill_zebra_1
                c_pval.alignment = align_center
                c_pval.border = border_all
                c_pval.font = Font(name='Calibri', size=10, color='555555')
            current_row += 1

    ws.column_dimensions['A'].width = 3 
    for col_idx in range(2, num_days + 2):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 11

    return wb