# app/modules/jobsistemi/routes.py

import os
import json
import platform
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill
from flask import Blueprint, jsonify, request

from app.globals import CONFIG_JOB_DIR
from app.core.helpers import load_json_config

jobsistemi_bp = Blueprint('jobsistemi_bp', __name__)

@jobsistemi_bp.route('/download-template', methods=['GET'])
def download_job_template():
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Configurazione Job"
        
        headers = ["Codice Azienda", "Nome Dipendente", "Codice Dipendente", "Causale in Excel", "Codice Causale Job"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="ea580c", end_color="ea580c", fill_type="solid")
        
        for row in range(2, 500):
            ws.cell(row=row, column=1).number_format = '@' 
            ws.cell(row=row, column=3).number_format = '@' 
            ws.cell(row=row, column=5).number_format = '@' 

        for col in range(1, 6):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 25

        home = Path.home()
        desktop_path = home / "Desktop"
        if platform.system() == 'Windows':
            onedrive_desktop = home / "OneDrive" / "Desktop"
            if onedrive_desktop.exists():
                desktop_path = onedrive_desktop
        
        file_path = desktop_path / "Inserimento Massivo JobSistemi.xlsx"
        wb.save(str(file_path))
        
        return jsonify({'success': True, 'message': 'Modello vuoto salvato con successo sul Desktop!'})
    except Exception as e:
        return jsonify({'success': False, 'message': str(e)})

@jobsistemi_bp.route('/delete-config', methods=['POST'])
def delete_job_config():
    data = request.json
    company = data.get('company')
    if not company: return jsonify({'success': False})
    path = os.path.join(CONFIG_JOB_DIR, f'{company}.json')
    if os.path.exists(path):
        os.remove(path)
    return jsonify({'success': True})

@jobsistemi_bp.route('/config', methods=['GET', 'POST'])
def handle_job_config():
    company = request.args.get('company')
    if request.method == 'GET':
        cfg = load_json_config(company, 'jobsistemi')
        if not cfg: cfg = {"company_code": "", "festivita_fisse": {}, "festivita_mobili": "", "employees": {}, "causals": {}}
        return jsonify({'success': True, 'config': cfg})
    else:
        data = request.json
        if not company: return jsonify({'success': False})
        path = os.path.join(CONFIG_JOB_DIR, f'{company}.json')
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data['config'], f, indent=4)
        return jsonify({'success': True})

@jobsistemi_bp.route('/import-config', methods=['POST'])
def import_job_config():
    try:
        file = request.files.get('file')
        if not file: return jsonify({'success': False, 'message': 'Nessun file caricato'})
        
        df = pd.read_excel(file)
        df.columns = [str(c).strip() for c in df.columns]
        df = df.fillna('')
        df = df.astype(str)
        
        config = {'company_code': '', 'festivita_fisse': {}, 'festivita_mobili': '', 'employees': {}, 'causals': {}}
        
        comp_col = next((c for c in df.columns if 'azienda' in c.lower()), None)
        emp_name_col = next((c for c in df.columns if 'nome' in c.lower() and 'dipendente' in c.lower()), None)
        emp_code_col = next((c for c in df.columns if 'codice' in c.lower() and 'dipendente' in c.lower()), None)
        caus_name_col = next((c for c in df.columns if 'causale' in c.lower() and 'excel' in c.lower()), None)
        if not caus_name_col:
            caus_name_col = next((c for c in df.columns if 'causale' in c.lower() and 'nome' not in c.lower()), None)
        caus_code_col = next((c for c in df.columns if 'codice' in c.lower() and 'causale' in c.lower()), None)

        first_valid_company = ""

        records = df.to_dict('records')
        for row in records:
            if not comp_col: continue
            
            comp_val = row.get(comp_col, '').strip()
            if comp_val == '' or comp_val.lower() == 'nan': continue
            
            row_comp_code = comp_val.split('.')[0].strip()
            if not first_valid_company: first_valid_company = row_comp_code

            if emp_name_col and emp_code_col:
                n = row.get(emp_name_col, '').strip()
                c = row.get(emp_code_col, '').strip()
                if n and c and n.lower() != 'nan' and c.lower() != 'nan':
                    config['employees'][n] = c.split('.')[0].strip()
                    
            if caus_name_col and caus_code_col:
                n = row.get(caus_name_col, '').strip()
                c = row.get(caus_code_col, '').strip()
                if n and c and n.lower() != 'nan' and c.lower() != 'nan':
                    config['causals'][n] = {'code': c.split('.')[0].strip(), 'festivity_action': 'ignora'}

        config['company_code'] = first_valid_company
        return jsonify({'success': True, 'config': config})
    except Exception as e:
        return jsonify({'success': False, 'message': f"Errore: {str(e)}"})